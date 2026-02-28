import openpyxl
import traceback
import time
import re
import os
from typing import Dict, List, Tuple
import asyncio
import threading  # Only retain progress bar lock, no longer used for LLM calls

import numpy as np
import pandas as pd
from tqdm.asyncio import tqdm_asyncio  # Asynchronous progress bar
from tqdm import tqdm  # Added: Synchronous progress bar
from requests.exceptions import RequestException, Timeout
from tenacity import (
    retry, stop_after_attempt, wait_exponential,
    retry_if_exception_type, AsyncRetrying,  # Asynchronous retry
)  # Synchronous retry is included in tenacity base imports

# Machine learning metrics library
from sklearn.metrics import precision_score, recall_score, f1_score, accuracy_score

# LangChain related (fix deprecation warnings + adapt to Python 3.12 asyncio)
from langchain_openai import ChatOpenAI
from langchain_core.prompts import PromptTemplate  # Replace langchain_classic
from pathlib import Path

# Remove RunnableSequence, use manual prompt splicing + LLM call (avoid context parameter issues)

# ===================== Core Configuration Area (Only adjust timeout-related) =====================
# Production-level call configuration (pure asynchronous version)
ASYNC_TIMEOUT = 300  # Total timeout for asynchronous tasks (replace original THREAD_TIMEOUT+API_TIMEOUT)
RETRY_ATTEMPTS = 3  # Number of retry attempts
RETRY_WAIT = (1, 5)  # Retry exponential backoff
semaphore = asyncio.Semaphore(5)  # Asynchronous rate limiting (effective)

# Other configurations (unchanged)
DATA_NAME = "哪吒之魔童闹海的影评"
DATA_ROOT_PATH = str(Path(__file__).parent.parent.parent.parent.absolute()) + "/data/"
# READING_FILE_LIST = [f"{DATA_NAME}_train_val_test_5_fold_{num}.xlsx" for num in range(1, 6)]
READING_FILE_LIST = [f"{DATA_NAME}_summary_version (deepseek-cot)_train_val_test_5_fold_{num}.xlsx" for num in range(1, 6)]

# RESULT_FILE = f"LLM_multi_label_w_cot_model_comparison_{DATA_NAME}_5fold_asyncio.xlsx"
RESULT_FILE = f"LLM_multi_label_w_cot_model_comparison_{DATA_NAME}_summary_version (deepseek-cot)_5fold_asyncio.xlsx"
FULL_SAVE_PATH = os.path.join(DATA_ROOT_PATH, "results_for_supervised_methods_full", RESULT_FILE)
EXPERIMENT_CONFIGS = [
    {"has_description": True, "has_few_shot": True, "name": "With Description + With Few-Shot"},
    {"has_description": True, "has_few_shot": False, "name": "With Description + Without Few-Shot"},
    {"has_description": False, "has_few_shot": True, "name": "Without Description + With Few-Shot"},
    {"has_description": False, "has_few_shot": False, "name": "Without Description + Without Few-Shot"},
]
# MODEL_LIST = ["qwen-plus", "qwen-flash", "deepseek-reasoner"]
MODEL_LIST = ["glm-4.5-flash", "glm-4.6", "qwen3-max", "qwen-plus", "qwen-flash", "deepseek-chat"]
# Added: Whether each model supports asynchronous calls (one-to-one with MODEL_LIST, users can adjust according to actual situation)
MODEL_ASYNC_SUPPORT = [False, False, True, True, True, True]
os.makedirs(os.path.dirname(FULL_SAVE_PATH), exist_ok=True)

# ===================== Global Variable Initialization (unchanged) =====================
aspects: List[str] = []
aspects_and_explanations: Dict[str, str] = {}
workbook = openpyxl.load_workbook(os.path.join(DATA_ROOT_PATH, "attributes_and_descriptions", f"aspects_{DATA_NAME}.xlsx"))
sheet = workbook.active
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    aspects.append(row[0])
    if row[1]:
        aspects_and_explanations[row[0]] = row[1]
workbook.close()

# 2. Load external prompt configuration (core modification: read prompt parts from xlsx)
prompt_config: Dict[str, str] = {}
unmentioned_label: int = 2  # Default unmentioned label (will be overwritten by file values)

try:
    # Load prompt configuration file
    prompt_file_path = os.path.join(DATA_ROOT_PATH, "prompts", f"prompt_{DATA_NAME}.xlsx")
    workbook_prompt = openpyxl.load_workbook(prompt_file_path)
    sheet_prompt = workbook_prompt.active

    # Read column names (ensure file column order: Task, Description, Rules, few_shot, Unmentioned_Label)
    # First row is column names, second row is configuration values
    prompt_rows = list(sheet_prompt.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    prompt_config = {
        "task": prompt_rows[0],  # Column 1: Core task description
        "description": prompt_rows[1],  # Column 2: Aspect description explanation
        "rules": prompt_rows[2],  # Column 3: Label rule explanation
        "few_shot": prompt_rows[3]  # Column 4: Example explanation
    }
    # Column 5: Unmentioned label value (convert to integer)
    unmentioned_label = int(prompt_rows[4]) if prompt_rows[4] is not None else 2
    workbook_prompt.close()
    print(f"✅ Successfully loaded prompt configuration | Unmentioned label value: {unmentioned_label}")
except Exception as e:
    print(f"⚠️ Failed to load prompt configuration, using default values | Exception: {str(e)}")
    # Default prompt configuration (compatible with original logic, can be deleted)
    prompt_config = {
        "task": '帮我分析评论"{review}"中关于多个方面"{aspects}"的情感，',
        "description": '其中每个方面的描述为：{description}。',
        "rules": """
                如果是正面的，回复1。
                如果是否定的，回复-1。
                如果评论与某个方面无关，回复0。
                由于方面有多个，结果用列表的方式给出：
                即类似于[1, -1, 0, ..., 0]的格式。
                其中每个元素的顺序与传入的方面顺序一一对应。
                仅返回最后的列表即可，不要返回其他内容。
                """,
        "few_shot": """
                如“纯粹念PPT，遇到深入一点的问题就直接跳过了。作为北大毕业生想申明我们的老师平时上课绝对不是这样的。”
                没有提到课程内容的逻辑性、覆盖面、趣味性和实用性，但说了纯粹念ppt显然是不好的教学方式，即教学方式有效性为负面，遇到深入一点的问题直接跳过，显示老师不具备专业的授课水平，因而老师的专业性为负面，上课只念ppt，表明老师上课态度敷衍，责任心不够，因此责任性为负面。
                因此最后的结果为[0, 0, 0, 0, -1, -1, -1]。
                再如“考研专业课帮助很大”，明显提到了课程对考研的帮助，即课程非常实用，其他方面不涉及。
                因此最后的结果为[0, 0, 0, 1, 0, 0, 0]。
                再如“非常有用，深入浅出”。
                其中，深入浅出显然表现了课程具备逻辑性，教学方法有效且老师授课比较专业，非常有用体现了课程的实用性，评论没有提及老师授课是否有责任心。
                因此最后的结果为[1, 0, 0, 1, 1, 1, 0]。
                """
    }


# ===================== Core Utility Functions (unchanged) =====================
def extract_last_int_list(text: str) -> List[int]:
    """Extract the last integer list from text (e.g., [1, -1, 0] from raw LLM response)"""
    pattern = r'.*\[(\s*[-+]?\d+\s*(?:,\s*[-+]?\d+\s*)*)\]'
    match = re.search(pattern, text, re.DOTALL)
    if not match:
        return []
    elements_str = match.group(1)
    int_elements = [int(elem.strip()) for elem in elements_str.split(',') if elem.strip()]
    return int_elements


def calculate_metrics(y_true: np.ndarray, y_pred: np.ndarray) -> Dict:
    """Calculate multi-label classification metrics (accuracy, precision, recall, F1)"""
    y_true_flat = y_true.flatten()
    y_pred_flat = y_pred.flatten()
    metrics = {
        'accuracy': accuracy_score(y_true_flat, y_pred_flat),
        'precision_macro': precision_score(y_true_flat, y_pred_flat, average='macro', zero_division=0),
        'recall_macro': recall_score(y_true_flat, y_pred_flat, average='macro', zero_division=0),
        'f1_macro': f1_score(y_true_flat, y_pred_flat, average='macro', zero_division=0),
        'precision_micro': precision_score(y_true_flat, y_pred_flat, average='micro', zero_division=0),
        'recall_micro': recall_score(y_true_flat, y_pred_flat, average='micro', zero_division=0),
        'f1_micro': f1_score(y_true_flat, y_pred_flat, average='micro', zero_division=0),
        'label_metrics': []
    }
    for label_idx in range(y_true.shape[1]):
        y_true_label = y_true[:, label_idx]
        y_pred_label = y_pred[:, label_idx]
        metrics['label_metrics'].append({
            'label': label_idx,
            'precision': precision_score(y_true_label, y_pred_label, average='macro', zero_division=0),
            'recall': recall_score(y_true_label, y_pred_label, average='macro', zero_division=0),
            'f1': f1_score(y_true_label, y_pred_label, average='macro', zero_division=0)
        })
    return metrics


def get_prompt_template(has_description: bool, has_few_shot: bool) -> PromptTemplate:
    """
    Dynamically build prompt template (core modification: read content from external configuration)
    :param has_description: Whether to include aspect description
    :param has_few_shot: Whether to include few-shot examples
    :return: Dynamically generated PromptTemplate
    """
    core_parts = [prompt_config["task"]]  # Column 1: Core task

    # Column 2: Aspect description (optional)
    if has_description and prompt_config["description"]:
        core_parts.append(prompt_config["description"])

    # Column 3: Label rules (required)
    core_parts.append(prompt_config["rules"])

    # Column 4: Few-shot examples (optional)
    if has_few_shot and prompt_config["few_shot"]:
        core_parts.append(prompt_config["few_shot"])

    # Splice and generate template
    full_template = "".join(core_parts)
    return PromptTemplate.from_template(full_template)


# ===================== Refactoring: Asynchronous + Synchronous LLM Calls (Added synchronous version) =====================
async def invoke_llm_safe_with_retry(
        llm: ChatOpenAI,
        prompt_template: PromptTemplate,
        review: str,
        review_idx: int,  # Added: Review index
        aspects: List[str],
        has_description: bool,
        aspects_and_explanations: Dict[str, str] = None
) -> Tuple[List[int], float]:
    """Asynchronous LLM call: adapt to dynamic unmentioned label + print failed review index"""
    start_time_single = time.time()
    async with semaphore:
        # Construct input data
        input_data = {
            'review': review,
            'aspects': ", ".join(aspects)
        }
        if has_description and aspects_and_explanations:
            desc_str = "\n".join([f"{k}: {v}" for k, v in aspects_and_explanations.items()])
            input_data['description'] = desc_str

        # Asynchronous retry logic
        async for attempt in AsyncRetrying(
                stop=stop_after_attempt(RETRY_ATTEMPTS),
                wait=wait_exponential(multiplier=1, min=RETRY_WAIT[0], max=RETRY_WAIT[1]),
                retry=retry_if_exception_type((RequestException, Timeout, Exception)),
                reraise=False
        ):
            with attempt:
                try:
                    # Manually render prompt and call LLM
                    prompt_str = prompt_template.format(**input_data)
                    resp = await asyncio.wait_for(
                        llm.ainvoke(prompt_str),
                        timeout=ASYNC_TIMEOUT
                    )
                    # Parse results
                    resp_text = resp.content
                    tags = extract_last_int_list(resp_text)[:len(aspects)]
                    # Fallback: fill with unmentioned label if tag length is insufficient (core modification)
                    tags = tags if len(tags) == len(aspects) else [unmentioned_label] * len(aspects)
                    elapsed_single = time.time() - start_time_single
                    return tags, elapsed_single
                except asyncio.TimeoutError:
                    print(f"\n⚠️ Timeout warning: Review index {review_idx} processing exceeded {ASYNC_TIMEOUT} seconds | Review snippet: {review[:50]}...")
                    elapsed_single = time.time() - start_time_single
                    return [unmentioned_label] * len(aspects), elapsed_single  # Dynamic unmentioned label
                except Exception as e:
                    if attempt.retry_state.attempt_number >= RETRY_ATTEMPTS:
                        print(
                            f"\n❌ Call failed: Review index {review_idx} failed after {RETRY_ATTEMPTS} retries | Exception: {str(e)[:100]} | Review snippet: {review[:50]}...")
                        traceback.print_exc()
                        elapsed_single = time.time() - start_time_single
                        return [unmentioned_label] * len(aspects), elapsed_single  # Dynamic unmentioned label
                    raise


def invoke_llm_safe_with_retry_sync(
        llm: ChatOpenAI,
        prompt_template: PromptTemplate,
        review: str,
        review_idx: int,  # Added: Review index
        aspects: List[str],
        has_description: bool,
        aspects_and_explanations: Dict[str, str] = None
) -> Tuple[List[int], float]:
    """Synchronous LLM call: adapt to dynamic unmentioned label + print failed review index (fixed fallback logic)"""
    start_time_single = time.time()
    # Construct input data
    input_data = {
        'review': review,
        'aspects': ", ".join(aspects)
    }
    if has_description and aspects_and_explanations:
        desc_str = "\n".join([f"{k}: {v}" for k, v in aspects_and_explanations.items()])
        input_data['description'] = desc_str

    # Synchronous retry logic
    @retry(
        stop=stop_after_attempt(RETRY_ATTEMPTS),
        wait=wait_exponential(multiplier=1, min=RETRY_WAIT[0], max=RETRY_WAIT[1]),
        retry=retry_if_exception_type((RequestException, Timeout, Exception)),
        reraise=False
    )
    def _sync_invoke():
        try:
            prompt_str = prompt_template.format(**input_data)
            # Synchronous LLM call (replace ainvoke with invoke)
            resp = llm.invoke(prompt_str)
            resp_text = resp.content
            tags = extract_last_int_list(resp_text)[:len(aspects)]
            # Fallback: fill with unmentioned label if tag length is insufficient (fixed: dynamically generated, no hardcoding)
            tags = tags if len(tags) == len(aspects) else [unmentioned_label] * len(aspects)
            return tags
        except TimeoutError:
            print(f"\n⚠️ Timeout warning: Review index {review_idx} processing exceeded {ASYNC_TIMEOUT} seconds | Review snippet: {review[:50]}...")
            return [unmentioned_label] * len(aspects)  # Dynamic unmentioned label
        except Exception as e:
            raise e

    try:
        tags = _sync_invoke()
    except Exception as e:
        print(f"\n❌ Synchronous call failed: Review index {review_idx} failed after {RETRY_ATTEMPTS} retries | Exception: {str(e)[:100]} | Review snippet: {review[:50]}...")
        traceback.print_exc()
        tags = [unmentioned_label] * len(aspects)  # Fixed: dynamically generate fallback labels

    elapsed_single = time.time() - start_time_single
    return tags, elapsed_single


# ===================== Single Fold Training Function (Adapt to asynchronous/synchronous models) =====================
async def train_single_fold(fold_number: int, file_path: str) -> Tuple[List[Dict], List[Dict]]:
    print(f"\n{'=' * 80}")
    print(f"Start processing Fold {fold_number} | File path: {file_path}")
    print(f"{'=' * 80}")

    # Read data (unchanged)
    try:
        excel_file = pd.ExcelFile(file_path)
        test_df = excel_file.parse('test', header=None)
        x_test = test_df.iloc[:, 0].astype(str).tolist()
        y_test = test_df.iloc[:, 1:].values
        y_test = np.array(y_test)
        excel_file.close()
    except Exception as e:
        print(f"❌ Failed to read Fold {fold_number} data: {str(e)}")
        return [], []

    fold_results = []
    fold_label_metrics = []

    # Load API keys (unchanged)
    api_key_glm = os.getenv("ZHIPU_API_KEY")
    api_key_tongyi = os.getenv("TONGYI_API_KEY")
    api_key_deepseek = os.getenv("DEEPSEEK_API_KEY")

    # Iterate over experiment configurations and models
    for exp_config in EXPERIMENT_CONFIGS:
        exp_name = exp_config["name"]
        has_description = exp_config["has_description"]
        has_few_shot = exp_config["has_few_shot"]
        print(f"\n--- Experiment Configuration: {exp_name} ---")

        # Pre-create prompt template (reusable)
        prompt_template = get_prompt_template(has_description, has_few_shot)

        # Iterate over models (added index to match async support list)
        for model_idx, model_name in enumerate(MODEL_LIST):
            is_async_supported = MODEL_ASYNC_SUPPORT[model_idx]
            print(f"\n----- Model: {model_name} (Async supported: {is_async_supported}) -----")

            # Initialize LLM (fix LangChain deprecation warnings + adapt to async/sync calls)
            try:
                if model_name.startswith("glm"):
                    llm = ChatOpenAI(
                        model=model_name,
                        api_key=api_key_glm,
                        base_url="https://open.bigmodel.cn/api/paas/v4/",
                        extra_body={"thinking": {"type": "disabled"}},
                        timeout=ASYNC_TIMEOUT  # Added LLM-level timeout
                    )
                elif model_name.startswith("deepseek"):
                    llm = ChatOpenAI(
                        model=model_name,
                        api_key=api_key_deepseek,
                        base_url="https://api.deepseek.com",
                        timeout=ASYNC_TIMEOUT  # Added LLM-level timeout
                    )
                elif model_name == "qwen3-max":
                    llm = ChatOpenAI(
                        model=model_name,
                        openai_api_key=api_key_tongyi,  # Please fill in your own APIKey
                        openai_api_base="https://dashscope.aliyuncs.com/compatible-mode/v1",
                        timeout=ASYNC_TIMEOUT  # Added LLM-level timeout
                    )
                else:  # Tongyi series
                    llm = ChatOpenAI(
                        model=model_name,
                        api_key=api_key_tongyi,
                        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
                        extra_body={"enable_thinking": False},
                        timeout=ASYNC_TIMEOUT  # Added LLM-level timeout
                    )
            except Exception as e:
                print(f"❌ Failed to initialize model {model_name}: {str(e)}, skipping this model")
                continue

            # Batch processing: distinguish async/sync logic
            start_time = time.time()
            if is_async_supported:
                # Asynchronous processing (original logic + pass review index)
                tasks = [
                    invoke_llm_safe_with_retry(
                        llm=llm,
                        prompt_template=prompt_template,
                        review=review,
                        review_idx=idx,  # Pass review index
                        aspects=aspects,
                        has_description=has_description,
                        aspects_and_explanations=aspects_and_explanations
                    ) for idx, review in enumerate(x_test)  # Enumerate to get index
                ]
                # Asynchronous progress bar (real-time update, task timeout will be terminated, progress bar will not get stuck)
                results = await tqdm_asyncio.gather(
                    *tasks,
                    desc=f"Processing reviews ({model_name} | {exp_name})",
                    total=len(tasks),
                    ncols=100
                )
                total_time = time.time() - start_time
                avg_time_per_sample_parallel = total_time / len(x_test)  # Parallel average time consumption
            else:
                # Synchronous processing (single thread + pass review index)
                results = []
                for idx, review in enumerate(tqdm(x_test, desc=f"Processing reviews ({model_name} | {exp_name} Sync)", ncols=100)):
                    tags, elapsed = invoke_llm_safe_with_retry_sync(
                        llm=llm,
                        prompt_template=prompt_template,
                        review=review,
                        review_idx=idx,  # Pass review index
                        aspects=aspects,
                        has_description=has_description,
                        aspects_and_explanations=aspects_and_explanations
                    )
                    results.append((tags, elapsed))
                total_time = time.time() - start_time
                avg_time_per_sample_parallel = '-'  # Async not supported, marked as None

            # Split label results and single sample time consumption (logic unchanged)
            y_pred = [r[0] for r in results]  # Original label prediction results
            single_times = [r[1] for r in results]  # Actual time consumption of each review (non-parallel)

            # Calculate metrics (unchanged)
            y_pred = np.array(y_pred)
            metrics = calculate_metrics(y_test, y_pred)

            # Calculate statistics of single sample time consumption (actual single sample processing time, not affected by parallelism)
            single_time_mean = np.mean(single_times)  # Average single sample time
            single_time_std = np.std(single_times)  # Standard deviation of single sample time
            single_time_min = np.min(single_times)  # Minimum single sample time
            single_time_max = np.max(single_times)  # Maximum single sample time

            # Print results (added async support annotation)
            print(f"\n📊 Fold {fold_number} | {exp_name} | {model_name} Results:")
            print(
                f"Accuracy: {metrics['accuracy']:.4f} | Macro F1: {metrics['f1_macro']:.4f} | Micro F1: {metrics['f1_micro']:.4f}")
            if is_async_supported:
                print(f"Parallel average time per sample: {avg_time_per_sample_parallel:.2f} seconds/sample")
            else:
                print(f"Parallel average time per sample: Async not supported")
            print(
                f"Actual average time per sample: {single_time_mean:.2f} seconds/sample (Std: {single_time_std:.2f}, Min: {single_time_min:.2f}, Max: {single_time_max:.2f})")

            # Save results (adapt to async/sync time statistics)
            fold_results.append({
                'fold_number': fold_number,
                'experiment_name': exp_name,
                'has_description': has_description,
                'has_few_shot': has_few_shot,
                'model_name': model_name,
                'is_async_supported': is_async_supported,  # Added: mark async support
                'accuracy': metrics['accuracy'],
                'precision_macro': metrics['precision_macro'],
                'recall_macro': metrics['recall_macro'],
                'f1_macro': metrics['f1_macro'],
                'precision_micro': metrics['precision_micro'],
                'recall_micro': metrics['recall_micro'],
                'f1_micro': metrics['f1_micro'],
                'avg_time_per_sample_parallel': avg_time_per_sample_parallel,  # Numeric for async, '-' for sync
                'avg_time_per_sample_single': single_time_mean,  # Actual average per sample
                'std_time_per_sample_single': single_time_std,  # Std of single sample time
                'min_time_per_sample_single': single_time_min,  # Min single sample time
                'max_time_per_sample_single': single_time_max  # Max single sample time
            })

            for label_metric in metrics['label_metrics']:
                fold_label_metrics.append({
                    'fold_number': fold_number,
                    'experiment_name': exp_name,
                    'has_description': has_description,
                    'has_few_shot': has_few_shot,
                    'model_name': model_name,
                    'is_async_supported': is_async_supported,  # Added: mark async support
                    'label_idx': label_metric['label'],
                    'label_name': aspects[label_metric['label']] if label_metric['label'] < len(aspects) else 'Unknown',
                    'precision': label_metric['precision'],
                    'recall': label_metric['recall'],
                    'f1': label_metric['f1']
                })

    return fold_results, fold_label_metrics


# ===================== 5-Fold Training Main Function (Changed to asynchronous) =====================
async def train_5folds() -> None:
    all_fold_results = []
    all_label_metrics = []

    for fold_idx, file_name in enumerate(READING_FILE_LIST, 1):
        file_path = os.path.join(DATA_ROOT_PATH, "reviews_after_split", file_name)
        if not os.path.exists(file_path):
            print(f"\n⚠️ Fold {fold_idx} file does not exist: {file_path}, skipping")
            continue
        # Asynchronous call to single fold training
        fold_result, fold_label_metric = await train_single_fold(fold_idx, file_path)
        all_fold_results.extend(fold_result)
        all_label_metrics.extend(fold_label_metric)

    if not all_fold_results:
        print("\n❌ No valid training results, terminate saving")
        return

    # Adjust summary statistics: exclude parallel time column with strings to avoid agg errors
    agg_funcs = {
        'accuracy': ['mean', 'std'],
        'precision_macro': ['mean', 'std'],
        'recall_macro': ['mean', 'std'],
        'f1_macro': ['mean', 'std'],
        'precision_micro': ['mean', 'std'],
        'recall_micro': ['mean', 'std'],
        'f1_micro': ['mean', 'std'],
        'avg_time_per_sample_single': ['mean', 'std'],
        'std_time_per_sample_single': ['mean', 'std'],
        'min_time_per_sample_single': ['min', 'max'],
        'max_time_per_sample_single': ['min', 'max']
    }
    df_folds = pd.DataFrame(all_fold_results)
    # Exclude parallel time column during group aggregation (avoid string cannot calculate mean/std)
    df_summary = df_folds.groupby(['experiment_name', 'model_name', 'is_async_supported']).agg(agg_funcs).round(4)
    df_summary.columns = ['_'.join(col).strip() for col in df_summary.columns.values]
    df_summary = df_summary.reset_index()
    df_labels = pd.DataFrame(all_label_metrics)

    with pd.ExcelWriter(FULL_SAVE_PATH, engine='openpyxl') as writer:
        df_folds.to_excel(writer, sheet_name='5-Fold Detailed Results', index=False)
        df_summary.to_excel(writer, sheet_name='Model + Experiment Summary', index=False)
        df_labels.to_excel(writer, sheet_name='Label-Level Detailed Metrics', index=False)

    print(f"\n{'=' * 80}")
    print(f"✅ 5-Fold Cross Validation completed! Results saved to: {FULL_SAVE_PATH}")
    print(f"📋 Excel contains 3 Sheets:")
    print(f"  1. 5-Fold Detailed Results: Raw metrics for each fold + experiment + model (with async support mark)")
    print(f"  2. Model + Experiment Summary: 5-fold mean + standard deviation (core comparison data)")
    print(f"  3. Label-Level Detailed Metrics: Precision/Recall/F1 for each aspect (with async support mark)")
    print(f"{'=' * 80}")


# ===================== Execution Entry (Asynchronous startup) =====================
if __name__ == "__main__":
    # Fixed: asyncio event loop issue for Python 3.8+ on Windows
    if os.name == 'nt':
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(train_5folds())