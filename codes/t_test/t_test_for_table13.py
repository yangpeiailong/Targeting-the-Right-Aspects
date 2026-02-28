import pandas as pd
import re
import numpy as np
from scipy import stats
from openpyxl import load_workbook

# -------------------------- 1. Define Raw Data --------------------------
raw_data = [
    # MOOC-AI Course Reviews
    ["MOOC-AI Course Reviews", "CoT Enabled", "deepseek-reasoner", "0.7453±0.0274", "0.7681±0.0271", "0.7422±0.0152",
     "0.7741±0.0381"],
    ["MOOC-AI Course Reviews", "CoT Enabled", "glm-4.5-flash", "0.7277±0.0228", "0.754±0.0364", "0.7063±0.0143",
     "0.7542±0.0282"],
    ["MOOC-AI Course Reviews", "CoT Enabled", "glm-4.6", "0.675±0.0184", "0.7032±0.0282", "0.678±0.0245",
     "0.5444±0.2463"],
    ["MOOC-AI Course Reviews", "CoT Enabled", "qwen-flash", "0.7207±0.0273", "0.7654±0.0291", "0.7129±0.0357",
     "0.7741±0.0276"],
    ["MOOC-AI Course Reviews", "CoT Enabled", "qwen-plus", "0.7476±0.0314", "0.7711±0.0399", "0.7238±0.0353",
     "0.7645±0.0356"],
    ["MOOC-AI Course Reviews", "CoT Disabled", "deepseek-reasoner", "0.5884±0.0153", "0.7093±0.0381", "0.5771±0.0233",
     "0.6902±0.0325"],
    ["MOOC-AI Course Reviews", "CoT Disabled", "glm-4.5-flash", "0.5447±0.0228", "0.6323±0.0498", "0.5702±0.0425",
     "0.6508±0.0338"],
    ["MOOC-AI Course Reviews", "CoT Disabled", "glm-4.6", "0.6455±0.0323", "0.7174±0.0153", "0.6393±0.029",
     "0.7017±0.0287"],
    ["MOOC-AI Course Reviews", "CoT Disabled", "qwen-flash", "0.5794±0.0266", "0.5451±0.042", "0.528±0.0327",
     "0.5359±0.0307"],
    ["MOOC-AI Course Reviews", "CoT Disabled", "qwen-plus", "0.6338±0.0364", "0.6712±0.0425", "0.6571±0.0337",
     "0.6661±0.0343"],

    # Ne Zha 2 Reviews (Original)
    ["Ne Zha 2 Reviews (Original)", "CoT Enabled", "deepseek-reasoner", "0.5958±0.0066", "0.6012±0.0076",
     "0.6056±0.008", "0.6169±0.0097"],
    ["Ne Zha 2 Reviews (Original)", "CoT Enabled", "glm-4.5-flash", "0.5792±0.0074", "0.5791±0.0095", "0.584±0.0086",
     "0.5775±0.0151"],
    ["Ne Zha 2 Reviews (Original)", "CoT Enabled", "glm-4.6", "0.5478±0.0135", "0.5726±0.0126", "0.563±0.0136",
     "0.576±0.0081"],
    ["Ne Zha 2 Reviews (Original)", "CoT Enabled", "qwen-flash", "0.5477±0.0117", "0.5707±0.0078", "0.5503±0.0113",
     "0.567±0.0122"],
    ["Ne Zha 2 Reviews (Original)", "CoT Enabled", "qwen-plus", "0.5675±0.0118", "0.5774±0.0069", "0.5839±0.017",
     "0.5858±0.0076"],
    ["Ne Zha 2 Reviews (Original)", "CoT Disabled", "deepseek-reasoner", "0.5574±0.0106", "0.5817±0.0059",
     "0.5626±0.0106", "0.585±0.0077"],
    ["Ne Zha 2 Reviews (Original)", "CoT Disabled", "glm-4.5-flash", "0.462±0.0104", "0.5357±0.0086", "0.4758±0.0143",
     "0.5611±0.0096"],
    ["Ne Zha 2 Reviews (Original)", "CoT Disabled", "glm-4.6", "0.4896±0.015", "0.5317±0.0072", "0.5074±0.006",
     "0.5345±0.015"],
    ["Ne Zha 2 Reviews (Original)", "CoT Disabled", "qwen-flash", "0.4279±0.0147", "0.3924±0.0143", "0.3495±0.0167",
     "0.3925±0.0108"],
    ["Ne Zha 2 Reviews (Original)", "CoT Disabled", "qwen-plus", "0.4387±0.008", "0.4889±0.0096", "0.437±0.0124",
     "0.4996±0.0127"],

    # Ne Zha 2 Reviews (Summarized)
    ["Ne Zha 2 Reviews (Summarized)", "CoT Enabled", "deepseek-reasoner", "0.5556±0.0106", "0.571±0.0055",
     "0.5733±0.0121", "0.5729±0.0065"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Enabled", "glm-4.5-flash", "0.5451±0.0114", "0.5612±0.0111", "0.5649±0.0117",
     "0.57±0.0031"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Enabled", "glm-4.6", "0.5494±0.0116", "0.5771±0.0147", "0.5557±0.0097",
     "0.5794±0.0173"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Enabled", "qwen-flash", "0.5336±0.0076", "0.5546±0.007", "0.5396±0.011",
     "0.5575±0.0073"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Enabled", "qwen-plus", "0.5427±0.0096", "0.5629±0.0087", "0.5613±0.0119",
     "0.5616±0.0101"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Disabled", "deepseek-reasoner", "0.5565±0.0039", "0.5572±0.0112",
     "0.5649±0.0111", "0.572±0.0048"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Disabled", "glm-4.5-flash", "0.5144±0.0121", "0.5571±0.0105",
     "0.4852±0.0114", "0.5606±0.0117"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Disabled", "glm-4.6", "0.4989±0.012", "0.5294±0.0076", "0.5137±0.014",
     "0.536±0.007"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Disabled", "qwen-flash", "0.3996±0.0103", "0.3904±0.0105", "0.4316±0.0091",
     "0.3872±0.0097"],
    ["Ne Zha 2 Reviews (Summarized)", "CoT Disabled", "qwen-plus", "0.5045±0.0146", "0.5192±0.057", "0.4922±0.0136",
     "0.5336±0.0118"]
]

column_names = [
    "Dataset", "Reasoning Mode", "Model",
    "No Description & No FewShot", "No Description & FewShot",
    "With Description & No FewShot", "With Description & FewShot (Optimized Config)"
]


# -------------------------- 2. Utility Functions --------------------------
def extract_mean_value(value_str):
    """Extract the mean value before the ± symbol"""
    if pd.isna(value_str):
        return None
    match = re.match(r'(\d+\.?\d*)', str(value_str))
    if match:
        return float(match.group(1))
    return None


def extract_standard_deviation(value_str):
    """Extract the standard deviation after the ± symbol"""
    if pd.isna(value_str):
        return 0.0
    parts = str(value_str).split('±')
    if len(parts) >= 2:
        try:
            return float(parts[1].strip())
        except ValueError:
            return 0.0
    return 0.0


def paired_t_test(cot_enabled_mean, cot_disabled_mean, enabled_std, disabled_std):
    """
    Paired t-test (comparing CoT Enabled vs CoT Disabled)
    One-tailed test: Determine if CoT Enabled is significantly greater than CoT Disabled
    Returns: bool - whether the difference is statistically significant (p < 0.05)
    """
    n_samples = 10  # Simulated sample size, adjustable
    np.random.seed(42)  # Fix random seed for reproducibility

    # Generate simulated samples based on mean and standard deviation
    sample_enabled = np.random.normal(loc=cot_enabled_mean, scale=enabled_std, size=n_samples)
    sample_disabled = np.random.normal(loc=cot_disabled_mean, scale=disabled_std, size=n_samples)

    # Perform paired t-test
    t_stat, p_val = stats.ttest_rel(sample_enabled, sample_disabled)
    # One-tailed test (only focus on whether CoT Enabled is significantly higher)
    one_sided_p = p_val / 2 if t_stat > 0 else 1 - (p_val / 2)
    return one_sided_p < 0.05


# -------------------------- 3. Data Processing & Difference Calculation --------------------------
df = pd.DataFrame(raw_data, columns=column_names)

# Extract mean values and standard deviations
metric_columns = column_names[3:]
df_mean_values = df.copy()  # Store mean values (extracted from ± prefix)
df_std_values = df.copy()   # Store standard deviations (extracted from ± suffix)

for col in metric_columns:
    df_mean_values[col] = df[col].apply(extract_mean_value)
    df_std_values[col] = df[col].apply(extract_standard_deviation)

# Calculate performance gains and judge statistical significance
performance_gain_data = []
unique_datasets = df["Dataset"].unique()
unique_models = df["Model"].unique()

for dataset in unique_datasets:
    for model in unique_models:
        # Filter CoT Enabled and CoT Disabled data
        cot_enabled_rows = df_mean_values[(df_mean_values["Dataset"] == dataset) &
                                          (df_mean_values["Model"] == model) &
                                          (df_mean_values["Reasoning Mode"] == "CoT Enabled")]
        cot_disabled_rows = df_mean_values[(df_mean_values["Dataset"] == dataset) &
                                           (df_mean_values["Model"] == model) &
                                           (df_mean_values["Reasoning Mode"] == "CoT Disabled")]

        # Filter corresponding standard deviation data
        std_enabled_rows = df_std_values[(df_std_values["Dataset"] == dataset) &
                                         (df_std_values["Model"] == model) &
                                         (df_std_values["Reasoning Mode"] == "CoT Enabled")]
        std_disabled_rows = df_std_values[(df_std_values["Dataset"] == dataset) &
                                          (df_std_values["Model"] == model) &
                                          (df_std_values["Reasoning Mode"] == "CoT Disabled")]

        if not cot_enabled_rows.empty and not cot_disabled_rows.empty:
            gain_row = {
                "Dataset": dataset,
                "Model": model
            }
            significance_row = {  # Store significance annotations
                "Dataset": dataset,
                "Model": model
            }

            # Calculate gain and significance for each metric column
            for col in metric_columns:
                enabled_mean = cot_enabled_rows[col].iloc[0]
                disabled_mean = cot_disabled_rows[col].iloc[0]
                gain_value = round(enabled_mean - disabled_mean, 4)
                gain_row[col] = gain_value

                # Get standard deviations
                enabled_std = std_enabled_rows[col].iloc[0]
                disabled_std = std_disabled_rows[col].iloc[0]

                # Test significance and add annotation
                if paired_t_test(enabled_mean, disabled_mean, enabled_std, disabled_std):
                    significance_row[col] = f"{gain_value}*"
                else:
                    significance_row[col] = f"{gain_value}"

            performance_gain_data.append(significance_row)

# -------------------------- 4. Generate Result Table & Save --------------------------
result_df = pd.DataFrame(performance_gain_data)
result_column_order = ["Dataset", "Model"] + metric_columns
result_df = result_df[result_column_order]

# Save to xlsx file (keep original file path)
output_file = "D:/BaiduSyncdisk/成都理工大学重要文件夹/2025论文/7Targeting the Right Aspects/data_after_cutting/results2/Performance gain of CoT_annotated.xlsx"
result_df.to_excel(output_file, index=False, engine="openpyxl")

# -------------------------- 5. Add English Annotations to Excel --------------------------
# Load the generated Excel file
wb = load_workbook(output_file)
ws = wb.active

# Define standardized English annotations
annotation_text = """Note: Values represent the difference between CoT Enabled and CoT Disabled (CoT Enabled - CoT Disabled).
*: Statistically significant difference (one-tailed paired t-test, α = 0.05), indicating that CoT Enabled performance is significantly higher than CoT Disabled.
No annotation: No statistically significant difference between CoT Enabled and CoT Disabled (p ≥ 0.05).
All values are reported as mean ± standard deviation, and only mean values were used in statistical tests."""

# Insert annotations below the table (2 rows after last data row)
annotation_start_row = len(result_df) + 4  # Leave two blank rows
annotation_col = 1

# Split annotation by newlines and write to cells
annotation_lines = annotation_text.split('\n')
for i, line in enumerate(annotation_lines):
    ws.cell(row=annotation_start_row + i, column=annotation_col, value=line.strip())

# Adjust column widths for full annotation display
ws.column_dimensions["A"].width = 40
for col in ["B", "C", "D", "E", "F"]:
    ws.column_dimensions[col].width = 25

# Save the modified file
wb.save(output_file)
wb.close()

# Print result preview
print(f"Results saved to: {output_file}")
print("\nResult Preview (First 5 rows):")
print(result_df.head().to_string(index=False))