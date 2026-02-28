import pandas as pd
import numpy as np
from scipy import stats
import openpyxl
from openpyxl import load_workbook
import os

# -------------------------- Configuration Parameters --------------------------
# File path (use raw string or escaped backslashes)
file_path = r"D:\BaiduSyncdisk\成都理工大学重要文件夹\2025论文\7Targeting the Right Aspects\data_after_cutting\results2\Core performance of LLMs under different configurations (F1-macro±std)_full.xlsx"
# Data range: Rows 2 to 34, Columns 4 to 7 (Excel column index starts at 1)
start_row = 2
end_row = 34
start_col = 4
end_col = 7


# -------------------------- Utility Functions --------------------------
def extract_main_value(value_str):
    """Extract the main value before ± from strings like '0.7453±0.0274'"""
    if pd.isna(value_str) or value_str == "":
        return np.nan
    # Split by ± symbol, take the first part and convert to float
    main_part = str(value_str).split('±')[0].strip()
    try:
        return float(main_part)
    except:
        return np.nan


def paired_t_test(max_val, other_val, std_max, std_other):
    """
    Paired t-test (simulate sample-level test based on mean and standard deviation)
    Assumes each value is the mean of multiple experiments, with std reflecting variability
    Returns True if statistically significant (p < 0.05)
    """
    # Simulate sample generation (default 10 samples, adjust as needed)
    n_samples = 10
    np.random.seed(42)  # Fix random seed for reproducibility
    sample_max = np.random.normal(loc=max_val, scale=std_max, size=n_samples)
    sample_other = np.random.normal(loc=other_val, scale=std_other, size=n_samples)

    # Paired t-test
    t_stat, p_val = stats.ttest_rel(sample_max, sample_other)
    # One-sided test (check if max_val is significantly greater than other_val)
    one_sided_p = p_val / 2 if t_stat > 0 else 1 - (p_val / 2)
    return one_sided_p < 0.05  # p<0.05 considered significant


def extract_std(value_str):
    """Extract the standard deviation after ± from strings like '0.7453±0.0274'"""
    if pd.isna(value_str) or value_str == "":
        return 0.0
    parts = str(value_str).split('±')
    if len(parts) >= 2:
        try:
            return float(parts[1].strip())
        except:
            return 0.0
    return 0.0


# -------------------------- Main Logic --------------------------
# 1. Read original Excel file (preserve all formatting)
wb = load_workbook(file_path)
ws = wb.active  # Get active worksheet

# 2. Process data by iterating through specified row range
for row_idx in range(start_row, end_row + 1):
    # Store original values, main values, and standard deviations for current row
    row_data = []
    main_values = []
    std_values = []

    # Read data from columns 4-7 of current row
    for col_idx in range(start_col, end_col + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_data.append(cell_value)
        main_val = extract_main_value(cell_value)
        std_val = extract_std(cell_value)
        main_values.append(main_val)
        std_values.append(std_val)

    # Skip rows containing null values or non-numeric values
    if any(np.isnan(v) for v in main_values):
        continue

    # Find index and value of the maximum value
    max_val = max(main_values)
    max_idx = main_values.index(max_val)
    max_std = std_values[max_idx]

    # Count significant results
    significant_count = 0
    # Perform t-test for each non-max value
    for i in range(len(main_values)):
        if i == max_idx:
            continue
        other_val = main_values[i]
        other_std = std_values[i]
        if paired_t_test(max_val, other_val, max_std, other_std):
            significant_count += 1

    # Determine annotation symbol
    if significant_count == 3:
        symbol = "*"  # All significant
    elif significant_count == 2:
        symbol = "++"  # Two significant
    elif significant_count == 1:
        symbol = "+"  # One significant
    else:
        symbol = ""  # No significance

    # Add annotation to the maximum value cell
    max_cell = ws.cell(row=row_idx, column=start_col + max_idx)
    original_value = max_cell.value
    if original_value and symbol:
        max_cell.value = f"{original_value}({symbol})"

# 3. Save file (overwrites original file - backup recommended first)
# Check if file is writable
if os.path.exists(file_path):
    try:
        wb.save(file_path)
        print(f"File saved successfully to: {file_path}")
    except PermissionError:
        # Save as new file if original is occupied
        new_file_path = file_path.replace(".xlsx", "_annotated.xlsx")
        wb.save(new_file_path)
        print(f"Original file is in use. Saved as new file: {new_file_path}")
else:
    wb.save(file_path)
    print(f"File saved to: {file_path}")

# Close workbook
wb.close()

# -------------------------- Validation Output --------------------------
# Read annotated data for verification
df_verify = pd.read_excel(file_path, header=0)
print("\nPreview of annotated data (Rows 2-5, Columns 4-7):")
print(df_verify.iloc[1:5, 3:7])