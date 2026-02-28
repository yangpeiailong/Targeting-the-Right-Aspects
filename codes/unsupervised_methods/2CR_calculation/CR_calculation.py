import openpyxl
import re
import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import warnings
from pathlib import Path
import os

warnings.filterwarnings('ignore')

# ===================== Configuration Parameters =====================
# Path to pre-trained Sentence-BERT model (Chinese fine-tuned)
BERT_MODEL_PATH = 'bert-base-chinese'
DATA_ROOT_PATH = str(Path(__file__).parent.parent.parent.absolute())
# Paths to predefined attribute files (key: dataset name)
ASPECTS_PATHS = {
    'AI_Course_Reviews': os.path.join(DATA_ROOT_PATH, "data/attributes_and_descriptions/aspects_人工智能1.xlsx"),
    'Nezha2_Movie_Reviews': os.path.join(DATA_ROOT_PATH, "data/attributes_and_descriptions/aspects_哪吒之魔童闹海的影评.xlsx")
}

# Paths to topic keyword tables (key: dataset_method)
TOPIC_TABLE_PATHS = {
    'AI_Course_Reviews_LDA': os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/人工智能1_LDA_8_Topics.xlsx"),
    'AI_Course_Reviews_KMeans': os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/人工智能1_KMeans_Clustering_Topics.xlsx"),
    'Nezha2_Movie_Reviews_LDA': os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/哪吒之魔童闹海的影评_LDA_5_Topics.xlsx"),
    'Nezha2_Movie_Reviews_KMeans': os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/哪吒之魔童闹海的影评_KMeans_Clustering_Topics.xlsx")
}

# Strict similarity threshold for short Chinese reviews (0.5 for discriminability)
SIMILARITY_THRESHOLD = 0.5

# Output path for four-dimensional metrics results
OUTPUT_PATH = os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/Four_Dimensional_CR_Results.xlsx")


# ======================================================================

def load_predefined_aspects(aspect_file_path):
    """
    Load predefined attributes from Excel file (only first column, ignore header/descriptions)

    Args:
        aspect_file_path (str): Path to Excel file with predefined aspects

    Returns:
        list[str]: Cleaned list of predefined attributes
    """
    workbook = openpyxl.load_workbook(aspect_file_path, read_only=True)
    worksheet = workbook.active
    predefined_aspects = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_index == 0:  # Skip header row
            continue
        aspect = row[0]
        if aspect and isinstance(aspect, str):  # Filter empty values
            predefined_aspects.append(aspect.strip())

    workbook.close()
    print(f"Loaded {len(predefined_aspects)} predefined aspects from {aspect_file_path}")
    return predefined_aspects


def parse_topic_keywords(topic_file_path, method_type):
    """
    Parse topic keywords from Excel files (support LDA/KMeans formats)
    - LDA: Column 1=Topic Number, Column 2=space-separated keywords
    - KMeans: Column 1=Topic ID, Column 3=keywords with similarity scores (cleaned)

    Args:
        topic_file_path (str): Path to Excel file with topic keywords
        method_type (str): Method type ('LDA'/'KMeans')

    Returns:
        dict: {topic_id: [keyword_1, keyword_2, ..., keyword_10]}
    """
    workbook = openpyxl.load_workbook(topic_file_path, read_only=True)
    worksheet = workbook.active
    topic_keywords_dict = {}

    # Regex to remove similarity scores (e.g., "monster(0.5778)" → "monster")
    similarity_pattern = re.compile(r'\([0-9.]+\)')

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_index == 0:  # Skip header row
            continue

        if method_type == 'LDA':
            topic_id = row[0]
            keywords_str = row[1] if row[1] else ''
            keywords = [kw.strip() for kw in keywords_str.split() if kw.strip()]

        elif method_type == 'KMeans':
            topic_id = row[0]
            keywords_str = row[2] if row[2] else ''
            clean_keywords_str = similarity_pattern.sub('', keywords_str)
            keywords = [kw.strip() for kw in clean_keywords_str.split() if kw.strip()]

        else:
            raise ValueError("method_type must be 'LDA' or 'KMeans'")

        if keywords:  # Filter empty topics
            topic_keywords_dict[topic_id] = keywords[:10]  # Keep top 10 keywords

    workbook.close()
    print(f"Parsed {len(topic_keywords_dict)} topics from {topic_file_path} ({method_type})")
    return topic_keywords_dict


def calculate_topic_vectors(topic_keywords_dict, embedding_model):
    """
    Calculate average semantic vector for each topic cluster

    Args:
        topic_keywords_dict (dict): {topic_id: [keywords]}
        embedding_model (SentenceTransformer): Pre-trained model

    Returns:
        dict: {topic_id: 768-dimensional cluster vector}
    """
    topic_vectors_dict = {}
    for topic_id, keywords in topic_keywords_dict.items():
        # Generate embeddings for keywords
        keyword_embeddings = embedding_model.encode(keywords, show_progress_bar=False)
        # Calculate average vector as topic center
        topic_vector = np.mean(keyword_embeddings, axis=0)
        topic_vectors_dict[topic_id] = topic_vector
    return topic_vectors_dict


def calculate_four_dimensional_metrics(aspects, topic_keywords_dict, embedding_model, threshold=0.8):
    """
    Calculate four-dimensional evaluation metrics: CR, MD, ACR, CE
    and return detailed coverage results

    Args:
        aspects (list[str]): Predefined attributes
        topic_keywords_dict (dict): {topic_id: [keywords]}
        embedding_model (SentenceTransformer): Pre-trained model
        threshold (float): Similarity threshold for effective coverage

    Returns:
        tuple: (detailed_results, metrics_dict)
    """
    # Step 1: Generate semantic vectors
    aspect_embeddings = embedding_model.encode(aspects, show_progress_bar=False)
    topic_vectors_dict = calculate_topic_vectors(topic_keywords_dict, embedding_model)

    # Step 2: Initialize detailed results storage
    detailed_results = {
        'aspect': [],
        'max_similarity': [],
        'matching_topic_id': [],
        'is_covered': []
    }

    # Step 3: Calculate similarity and coverage for each attribute
    total_attributes = len(aspects)
    total_topics = len(topic_keywords_dict)
    contributing_topics = set()  # Topics that cover at least one attribute

    for idx, aspect in enumerate(aspects):
        aspect_vector = aspect_embeddings[idx].reshape(1, -1)
        max_sim = 0.0
        matching_topic_id = None

        # Find topic with maximum similarity
        for topic_id, topic_vector in topic_vectors_dict.items():
            topic_vector = topic_vector.reshape(1, -1)
            sim = cosine_similarity(aspect_vector, topic_vector)[0][0]
            if sim > max_sim:
                max_sim = sim
                matching_topic_id = topic_id

        # Record detailed results
        is_covered = max_sim >= threshold
        detailed_results['aspect'].append(aspect)
        detailed_results['max_similarity'].append(round(max_sim, 6))
        detailed_results['matching_topic_id'].append(matching_topic_id)
        detailed_results['is_covered'].append(is_covered)

        # Add to contributing topics if covered
        if is_covered and matching_topic_id is not None:
            contributing_topics.add(matching_topic_id)

    # Step 4: Calculate four-dimensional metrics
    covered_attributes = sum(detailed_results['is_covered'])
    num_contributing_topics = len(contributing_topics)

    # 1. Core Coverage Rate (CR)
    cr = (covered_attributes / total_attributes) * 100 if total_attributes > 0 else 0.0

    # 2. Mapping Density (MD)
    md = covered_attributes / total_topics if total_topics > 0 else 0.0

    # 3. Adjusted Coverage Rate (ACR)
    topic_attr_ratio = total_topics / total_attributes if total_attributes > 0 else 0.0
    acr = cr * min(1.0, topic_attr_ratio)

    # 4. Coverage Evenness (CE)
    ce = (num_contributing_topics / total_topics) * 100 if total_topics > 0 else 0.0

    # Compile metrics dictionary
    metrics_dict = {
        'cr': round(cr, 2),
        'md': round(md, 2),
        'acr': round(acr, 2),
        'ce': round(ce, 2),
        'total_attributes': total_attributes,
        'covered_attributes': covered_attributes,
        'total_topics': total_topics,
        'num_contributing_topics': num_contributing_topics
    }

    # Print summary for verification
    print("\n=== Four-Dimensional Metrics Summary ===")
    print(f"Total Attributes: {total_attributes} | Covered: {covered_attributes}")
    print(f"Total Topics: {total_topics} | Contributing Topics: {num_contributing_topics}")
    print(f"CR: {metrics_dict['cr']}% | MD: {metrics_dict['md']}")
    print(f"ACR: {metrics_dict['acr']}% | CE: {metrics_dict['ce']}%")

    return detailed_results, metrics_dict


def save_four_dimensional_results(all_results, output_path):
    """
    Save detailed coverage results and four-dimensional metrics to Excel
    - Separate worksheets for detailed results (per dataset-method)
    - Summary worksheet for all metrics

    Args:
        all_results (dict): Aggregated results for all dataset-method combinations
        output_path (str): Path to save Excel file
    """
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # 1. Save detailed coverage results (per dataset-method)
    for key, data in all_results.items():
        # Create worksheet for detailed results
        ws_detail = workbook.create_sheet(title=f"{key}_Detailed")
        # Write header
        ws_detail.append(['Predefined Aspect', 'Max Similarity', 'Matching Topic ID', 'Is Covered (≥{threshold})'.format(threshold=SIMILARITY_THRESHOLD)])
        # Write detailed data
        for idx in range(len(data['detailed']['aspect'])):
            ws_detail.append([
                data['detailed']['aspect'][idx],
                data['detailed']['max_similarity'][idx],
                data['detailed']['matching_topic_id'][idx],
                data['detailed']['is_covered'][idx]
            ])

    # 2. Save four-dimensional metrics summary (core results for paper)
    ws_summary = workbook.create_sheet(title='Four_Dimensional_Summary', index=0)
    # Write header
    ws_summary.append([
        'Dataset', 'Method', 'Total Attributes', 'Covered Attributes',
        'Total Topics', 'Contributing Topics', 'CR (%)', 'MD', 'ACR (%)', 'CE (%)'
    ])

    # Write metrics data
    for key, data in all_results.items():
        # Split dataset and method from key (handle multi-underscore dataset names)
        method = key.split('_')[-1]
        dataset = '_'.join(key.split('_')[:-1])
        metrics = data['metrics']

        ws_summary.append([
            dataset,
            method,
            metrics['total_attributes'],
            metrics['covered_attributes'],
            metrics['total_topics'],
            metrics['num_contributing_topics'],
            metrics['cr'],
            metrics['md'],
            metrics['acr'],
            metrics['ce']
        ])

    # Save Excel file
    workbook.save(output_path)
    print(f"\nAll four-dimensional results saved to: {output_path}")


def main():
    """Main pipeline for four-dimensional coverage metrics calculation"""
    # Load pre-trained Sentence-BERT model
    print("Loading Sentence-BERT model (Chinese fine-tuned)...")
    embedding_model = SentenceTransformer(BERT_MODEL_PATH)

    # Initialize results storage
    all_results = {}

    # Process all dataset-method combinations
    for dataset_key in ['AI_Course_Reviews', 'Nezha2_Movie_Reviews']:
        # Load predefined attributes for current dataset
        aspects = load_predefined_aspects(ASPECTS_PATHS[dataset_key])

        # Process LDA results
        lda_key = f"{dataset_key}_LDA"
        if lda_key in TOPIC_TABLE_PATHS:
            print(f"\n===== Processing {lda_key} =====")
            topic_keywords = parse_topic_keywords(TOPIC_TABLE_PATHS[lda_key], method_type='LDA')
            detailed, metrics = calculate_four_dimensional_metrics(
                aspects, topic_keywords, embedding_model, SIMILARITY_THRESHOLD
            )
            all_results[lda_key] = {'detailed': detailed, 'metrics': metrics}

        # Process KMeans results
        kmeans_key = f"{dataset_key}_KMeans"
        if kmeans_key in TOPIC_TABLE_PATHS:
            print(f"\n===== Processing {kmeans_key} =====")
            topic_keywords = parse_topic_keywords(TOPIC_TABLE_PATHS[kmeans_key], method_type='KMeans')
            detailed, metrics = calculate_four_dimensional_metrics(
                aspects, topic_keywords, embedding_model, SIMILARITY_THRESHOLD
            )
            all_results[kmeans_key] = {'detailed': detailed, 'metrics': metrics}

    # Save all results to Excel
    save_four_dimensional_results(all_results, OUTPUT_PATH)
    print("\n===== All Four-Dimensional Metrics Calculations Completed =====")


if __name__ == '__main__':
    main()