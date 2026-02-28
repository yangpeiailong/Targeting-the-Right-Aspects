import openpyxl
import re
import jieba
import jieba.posseg as pseg
import warnings
import numpy as np
from openpyxl import Workbook
from fasttext import load_model
from pathlib import Path
import os

# Ignore irrelevant warnings
warnings.filterwarnings('ignore', category=UserWarning, module='pkg_resources')
warnings.filterwarnings('ignore', category=UserWarning, module='jieba')
warnings.filterwarnings('ignore')

# ===================== Configurable Parameters =====================
DATASET_NAME = '哪吒之魔童闹海的影评'  # Keep Chinese for file path compatibility
DATA_ROOT_PATH = str(Path(__file__).parent.parent.parent.absolute())
excel_file_path = os.path.join(DATA_ROOT_PATH, f"data/reviews_after_annotation/{DATASET_NAME}_attributes_for_each_review2.xlsx")
stopword_file_path = os.path.join(DATA_ROOT_PATH, f"data/stop_words/stop_words.txt")
fasttext_model_path = 'D:/fasttext/cc.zh.300.bin'
vector_output_path = os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/{DATASET_NAME}_Nouns_WordVectors.xlsx")

# ===================================================================

def filter_special_characters(text):
    """
    Filter special characters, non-Chinese characters, numbers and single-character symbols
    Args:
        text (str): Original text string
    Returns:
        str: Cleaned text with only Chinese characters
    """
    text = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', text)
    text = re.sub(r'[a-zA-Z]', '', text)
    text = re.sub(r'\d+', '', text)
    return text

def load_and_preprocess_text_data():
    """
    Load and preprocess text data, extract noun vocabulary with POS tagging and special char filtering
    Returns:
        valid_nouns (list): List of unique valid nouns (POS starts with 'n')
    """
    # Load Excel file and read text content
    workbook = openpyxl.load_workbook(excel_source_path, read_only=True)
    worksheet = workbook.active

    raw_texts = []
    word_frequency = {}
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_index == 0:
            continue
        first_column_content = row[0]
        if not first_column_content or not isinstance(first_column_content, str):
            continue

        # Filter special characters before POS tagging
        cleaned_text = filter_special_characters(first_column_content)
        if not cleaned_text:
            continue

        # POS tagging and frequency calculation
        tokenized_text_with_pos = pseg.lcut(cleaned_text)
        tokenized_text = []
        for word, pos in tokenized_text_with_pos:
            stripped_word = word.strip()
            if stripped_word:
                tokenized_text.append(stripped_word)
                word_frequency[stripped_word] = word_frequency.get(stripped_word, 0) + 1
        raw_texts.append(tokenized_text)

    # Load stopwords with encoding compatibility
    try:
        with open(stopword_file_path, 'r', encoding='gbk') as f:
            stopwords = [word.strip() for word in f.readlines()]
    except UnicodeDecodeError:
        with open(stopword_file_path, 'r', encoding='utf-8') as f:
            stopwords = [word.strip() for word in f.readlines()]

    # Filter nouns with POS tagging
    print("\nFiltering nouns (POS tags starting with 'n')...")
    noun_frequency = {}
    workbook = openpyxl.load_workbook(excel_source_path, read_only=True)
    worksheet = workbook.active
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_index == 0:
            continue
        first_column_content = row[0]
        if not first_column_content or not isinstance(first_column_content, str):
            continue

        # Filter special characters before POS tagging
        cleaned_text = filter_special_characters(first_column_content)
        if not cleaned_text:
            continue

        # POS tagging for noun filtering
        tokenized_text_with_pos = pseg.lcut(cleaned_text)
        for word, pos in tokenized_text_with_pos:
            stripped_word = word.strip()
            # Filter valid nouns with multiple conditions
            if (pos.startswith('n') and
                    stripped_word not in stopwords and
                    len(stripped_word) > 1 and
                    re.match(r'^[\u4e00-\u9fa5]+$', stripped_word)):
                noun_frequency[stripped_word] = noun_frequency.get(stripped_word, 0) + 1

    # Extract unique valid nouns
    valid_nouns = list(noun_frequency.keys())

    # Print statistical information
    print(f"Total number of unique valid words (all types): {len(word_frequency)}")
    print(f"Total number of filtered nouns: {len(valid_nouns)}")
    print(f"Total occurrence count of nouns (including duplicates): {sum(noun_frequency.values())}")
    # Print top 10 frequent nouns
    sorted_nouns = sorted(noun_frequency.items(), key=lambda x: x[1], reverse=True)
    if sorted_nouns:
        print(f"Top 10 frequent nouns: {[word for word, freq in sorted_nouns[:10]]}")

    return valid_nouns

def generate_fasttext_vectors(words):
    """
    Load FastText model and generate word vectors for input words
    Args:
        words (list): List of words to generate vectors for
    Returns:
        word_vectors (dict): Key: word, Value: 300-dimensional vector
    """
    print(f"\nLoading FastText model from: {fasttext_model_path}")
    fasttext_model = load_model(fasttext_model_path)

    # Generate word vectors and filter OOV words
    word_vectors = {}
    oov_words = []
    for word in words:
        try:
            vector = fasttext_model.get_word_vector(word)
            word_vectors[word] = vector
        except Exception as e:
            oov_words.append(word)

    # Print vector generation results
    print(f"Number of nouns with successfully generated vectors: {len(word_vectors)}")
    print(f"Number of out-of-vocabulary (OOV) words: {len(oov_words)}")
    if len(oov_words) > 0:
        print(f"Example OOV words: {oov_words[:10]}")

    return word_vectors

def save_word_vectors_to_excel(word_vectors):
    """
    Save nouns and their corresponding word vectors to Excel file
    Args:
        word_vectors (dict): Key: noun, Value: 300-dimensional vector
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Nouns-WordVectors"

    # Set header row
    worksheet['A1'] = 'Vocabulary'
    for dimension in range(300):
        worksheet.cell(row=1, column=dimension + 2, value=f"Dimension {dimension + 1}")

    # Write word vectors to Excel
    row_index = 2
    for word, vector in word_vectors.items():
        worksheet.cell(row=row_index, column=1, value=word)
        for col_index, value in enumerate(vector, start=2):
            worksheet.cell(row=row_index, column=col_index, value=float(value))
        row_index += 1

    # Save Excel file
    workbook.save(vector_output_path)
    print(f"\nNoun word vectors saved to: {vector_output_path}")

if __name__ == '__main__':
    valid_nouns = load_and_preprocess_text_data()
    word_vectors = generate_fasttext_vectors(valid_nouns)
    save_word_vectors_to_excel(word_vectors)