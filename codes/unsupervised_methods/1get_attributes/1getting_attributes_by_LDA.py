import openpyxl
from gensim import corpora, models
import jieba
import warnings
from openpyxl import Workbook
import re
from pathlib import Path
import os

# Ignore irrelevant warnings
warnings.filterwarnings('ignore', category=UserWarning, module='pkg_resources')
warnings.filterwarnings('ignore', category=UserWarning, module='jieba')
warnings.filterwarnings('ignore')

# ===================== Configurable Parameters =====================
DATASET_NAME = '人工智能1'  # Keep Chinese for file path compatibility
DATA_ROOT_PATH = str(Path(__file__).parent.parent.parent.absolute())
excel_file_path = os.path.join(DATA_ROOT_PATH, f"data/reviews_after_annotation/{DATASET_NAME}_attributes_for_each_review2.xlsx")
stopword_file_path = os.path.join(DATA_ROOT_PATH, f"data/stop_words/stop_words.txt")
OPTIMAL_TOPIC_COUNT = 8
TOPIC_KEYWORD_COUNT = 10
excel_output_path = os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/{DATASET_NAME}_LDA_{OPTIMAL_TOPIC_COUNT}_Topics.xlsx")
RANDOM_SEED = 42

# ===================================================================

def filter_special_characters(text):
    """
    Filter special characters, non-Chinese characters, numbers and single-character symbols
    Args:
        text (str): Original text string
    Returns:
        str: Cleaned text with only Chinese characters
    """
    text = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', text)
    text = re.sub(r'[a-zA-Z]', '', text)
    text = re.sub(r'\d+', '', text)
    return text

def load_and_preprocess_text_data():
    """
    Load and preprocess text data from Excel file with special character filtering
    Returns:
        processed_texts (list): Tokenized texts after stopword removal and special char filtering
        text_dictionary (gensim.corpora.Dictionary): Gensim dictionary object
        bow_corpus (list): Bag-of-words corpus
        tfidf_corpus (list): TF-IDF weighted corpus
    """
    workbook = openpyxl.load_workbook(excel_source_path, read_only=True)
    worksheet = workbook.active

    raw_texts = []
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_index == 0:
            continue
        first_column_content = row[0]
        if not first_column_content or not isinstance(first_column_content, str):
            continue

        cleaned_text = filter_special_characters(first_column_content)
        tokenized_text = [token.strip() for token in jieba.lcut(cleaned_text) if token.strip()]
        raw_texts.append(tokenized_text)

    # Load stopwords with encoding compatibility
    try:
        with open(stopword_file_path, 'r', encoding='gbk') as f:
            stopwords = [word.strip() for word in f.readlines()]
    except UnicodeDecodeError:
        with open(stopword_file_path, 'r', encoding='utf-8') as f:
            stopwords = [word.strip() for word in f.readlines()]

    # Filter stopwords, empty strings and non-Chinese single characters
    processed_texts = []
    for text in raw_texts:
        filtered_tokens = []
        for token in text:
            if (token not in stopwords and
                token and
                len(token) > 1 and
                re.match(r'^[\u4e00-\u9fa5]+$', token)):
                filtered_tokens.append(token)
        if filtered_tokens:
            processed_texts.append(filtered_tokens)

    # Create dictionary and corpus with extreme frequency filtering
    text_dictionary = corpora.Dictionary(processed_texts)
    text_dictionary.filter_extremes(no_below=2, no_above=0.9)
    bow_corpus = [text_dictionary.doc2bow(text) for text in processed_texts]

    # Convert to TF-IDF representation
    tfidf_transformer = models.TfidfModel(bow_corpus)
    tfidf_corpus = tfidf_transformer[bow_corpus]

    print(f"Preprocessing completed: {len(processed_texts)} valid texts, {len(text_dictionary)} unique tokens")
    return processed_texts, text_dictionary, bow_corpus, tfidf_corpus

def train_final_lda_model(text_dictionary, tfidf_corpus):
    """
    Train the final LDA model with optimal number of topics
    Args:
        text_dictionary (gensim.corpora.Dictionary): Gensim dictionary
        tfidf_corpus (list): TF-IDF weighted corpus
    Returns:
        lda_model (gensim.models.LdaModel): Trained LDA model
    """
    print(f"\nTraining LDA model with {OPTIMAL_TOPIC_COUNT} topics...")
    lda_model = models.LdaModel(
        corpus=tfidf_corpus,
        id2word=text_dictionary,
        num_topics=OPTIMAL_TOPIC_COUNT,
        random_state=RANDOM_SEED,
        update_every=1,
        chunksize=100,
        passes=10,
        alpha='auto',
        per_word_topics=True
    )
    print("LDA model training completed!")
    return lda_model

def extract_topic_keywords(lda_model):
    """
    Extract topic keywords and format for Excel output
    Args:
        lda_model (gensim.models.LdaModel): Trained LDA model
    Returns:
        topic_data (list): List of [topic_number, keyword_string] for Excel export
    """
    topic_data = []
    for topic_id in range(OPTIMAL_TOPIC_COUNT):
        topic_keywords = lda_model.show_topic(topic_id, topn=TOPIC_KEYWORD_COUNT)
        keywords_list = [word for word, weight in topic_keywords]
        keywords_string = ' '.join(keywords_list)
        topic_data.append([topic_id + 1, keywords_string])
    return topic_data

def save_topics_to_excel(topic_data):
    """
    Save extracted topic data to Excel file
    Args:
        topic_data (list): List of [topic_number, keyword_string]
    """
    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = f"{OPTIMAL_TOPIC_COUNT} Topics"

    # Set header
    output_worksheet['A1'] = 'Topic Number'
    output_worksheet['B1'] = 'Topic Keywords'

    # Write topic data to Excel
    for row_index, (topic_number, keywords) in enumerate(topic_data, start=2):
        output_worksheet.cell(row=row_index, column=1, value=topic_number)
        output_worksheet.cell(row=row_index, column=2, value=keywords)

    # Adjust column width
    output_worksheet.column_dimensions['A'].width = 10
    output_worksheet.column_dimensions['B'].width = 80

    # Save Excel file
    output_workbook.save(excel_output_path)
    print(f"\nTopics successfully saved to Excel file: {excel_output_path}")

if __name__ == '__main__':
    processed_texts, text_dict, bow_corpus, tfidf_corpus = load_and_preprocess_text_data()
    lda_model = train_final_lda_model(text_dict, tfidf_corpus)
    topic_data = extract_topic_keywords(lda_model)
    save_topics_to_excel(topic_data)

    # Print topic preview
    print("\n=== Topic Preview ===")
    for topic_number, keywords in topic_data:
        print(f"Topic {topic_number}: {keywords}")