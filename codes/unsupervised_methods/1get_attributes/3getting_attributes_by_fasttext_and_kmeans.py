import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from sklearn.cluster import KMeans
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.preprocessing import normalize
from fasttext import load_model
from pathlib import Path
import os

# Ignore irrelevant warnings
warnings.filterwarnings('ignore')

# ===================== Configuration Parameters =====================
DATASET_NAME = '哪吒之魔童闹海的影评'  # Keep Chinese for file path compatibility
DATA_ROOT_PATH = str(Path(__file__).parent.parent.parent.absolute())
# Path to input Excel file containing word vectors (complete path)
vector_input_path = os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/{DATASET_NAME}_Nouns_WordVectors.xlsx")
# Path to pre-trained FastText model (verify path correctness)
fasttext_model_path = 'D:/fasttext/cc.zh.300.bin'
# Optimal number of clusters (derived from previous KMeans analysis script)
OPTIMAL_CLUSTER_COUNT = 16
# Number of top words to output per topic (consistent with LDA/AP clustering)
TOP_N_KEYWORDS = 10
# Path to save KMeans clustering topic results (complete path)
topic_output_path = os.path.join(DATA_ROOT_PATH, f"data/results_for_unsupervised_methods/{DATASET_NAME}_KMeans_Clustering_Topics.xlsx")

# ======================================================================

def load_word_vectors_from_excel():
    """
    Load vocabulary and corresponding word vectors from Excel file
    Returns:
        words (list): List of valid vocabulary words
        normalized_vectors (numpy.ndarray): L2-normalized word vectors (shape: [n_words, 300])
    """
    print(f"Loading word vector file: {vector_input_path}")
    df = pd.read_excel(vector_input_path)

    # Extract vocabulary and vectors from DataFrame
    words = df['Vocabulary'].tolist()
    vectors = df.iloc[:, 1:].values  # Vectors start from the second column

    # Filter out invalid vectors (containing NaN values)
    valid_indices = ~np.isnan(vectors).any(axis=1)
    words = [words[i] for i in range(len(words)) if valid_indices[i]]
    vectors = vectors[valid_indices]

    # L2 normalization to improve KMeans clustering performance
    normalized_vectors = normalize(vectors, norm='l2')

    print(f"Loading completed. Number of valid words: {len(words)}")
    return words, normalized_vectors


def perform_kmeans_clustering(words, vectors):
    """
    Perform KMeans clustering with optimal number of clusters
    Args:
        words (list): List of valid vocabulary words
        vectors (numpy.ndarray): Normalized word vectors
    Returns:
        clusters (dict): Key: cluster ID, Value: list of words in cluster
        cluster_centers (numpy.ndarray): Cluster center vectors (shape: [n_clusters, 300])
        vectors (numpy.ndarray): Original normalized vectors
        words (list): Original list of valid words
    """
    print(f"\nPerforming KMeans clustering (Number of clusters: {OPTIMAL_CLUSTER_COUNT})...")

    # Initialize KMeans model with optimal parameters
    kmeans_model = KMeans(
        n_clusters=OPTIMAL_CLUSTER_COUNT,
        random_state=42,
        n_init='auto'
    )
    cluster_labels = kmeans_model.fit_predict(vectors)

    # Organize clustering results: {cluster_id: [word1, word2, ...]}
    clusters = {}
    for index, label in enumerate(cluster_labels):
        if label not in clusters:
            clusters[label] = []
        clusters[label].append(words[index])

    # Calculate cluster center vectors (for subsequent similarity sorting)
    cluster_centers = kmeans_model.cluster_centers_

    # Print number of words per cluster (verify clustering effectiveness)
    print("\nNumber of words per cluster:")
    for cluster_id in sorted(clusters.keys()):
        print(f"Cluster {cluster_id + 1}: {len(clusters[cluster_id])} words")

    print(f"\nKMeans clustering completed. Generated {len(clusters)} topics (clusters)")
    return clusters, cluster_centers, vectors, words


def generate_topic_keywords(clusters, cluster_centers, vectors, words, fasttext_model):
    """
    Generate topic keywords sorted by cosine similarity to cluster center
    Args:
        clusters (dict): Organized KMeans clustering results
        cluster_centers (numpy.ndarray): Cluster center vectors
        vectors (numpy.ndarray): Normalized word vectors
        words (list): List of valid vocabulary words
        fasttext_model (fasttext.FastText._FastText): Loaded FastText model (unused, kept for compatibility)
    Returns:
        topics (list): List of [topic_id, center_word, keyword_string] for Excel export
    """
    topics = []

    print("\nGenerating topic keywords (sorted by cosine similarity to cluster center):")
    for cluster_id in sorted(clusters.keys()):
        # Get all words in current cluster
        cluster_words = clusters[cluster_id]
        # Get center vector of current cluster
        center_vector = cluster_centers[cluster_id]

        # Find indices of current cluster words in the full word list
        cluster_word_indices = [words.index(word) for word in cluster_words]
        # Extract vectors for current cluster words
        cluster_vectors = vectors[cluster_word_indices]

        # Calculate cosine similarity between each word and cluster center
        similarities = cosine_similarity(cluster_vectors, center_vector.reshape(1, -1)).flatten()

        # Sort words by similarity (descending order)
        word_similarity_pairs = list(zip(cluster_words, similarities))
        word_similarity_pairs.sort(key=lambda x: x[1], reverse=True)

        # Select top N keywords
        top_keywords = word_similarity_pairs[:TOP_N_KEYWORDS]
        # Cluster center word (word with highest similarity)
        center_word = top_keywords[0][0]

        # Format output: word(similarity_value)
        keyword_string = ' '.join([f"{word}({sim:.4f})" for word, sim in top_keywords])
        # Topic ID starts from 1 (consistent with LDA/AP clustering output)
        topics.append([cluster_id + 1, center_word, keyword_string])

        print(f"Topic {cluster_id + 1} (Center: {center_word}): {keyword_string}")

    return topics


def save_topic_results_to_excel(topics):
    """
    Save KMeans clustering topic results to Excel file (aligned with LDA/AP format)
    Args:
        topics (list): List of [topic_id, center_word, keyword_string]
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "KMeans_Clustering_Topics"

    # Set header (consistent with AP clustering output)
    worksheet['A1'] = 'Topic ID'
    worksheet['B1'] = 'Cluster Center Word'
    worksheet['C1'] = 'Topic Keywords (with Similarity)'

    # Write data to Excel
    for row_index, (topic_id, center_word, keywords) in enumerate(topics, start=2):
        worksheet.cell(row=row_index, column=1, value=topic_id)
        worksheet.cell(row=row_index, column=2, value=center_word)
        worksheet.cell(row=row_index, column=3, value=keywords)

    # Adjust column widths for better readability
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 100

    # Save Excel file
    workbook.save(topic_output_path)
    print(f"\nTopic results saved to: {topic_output_path}")


if __name__ == '__main__':
    # Step 1: Load vocabulary and normalized word vectors
    vocabulary_words, normalized_vectors = load_word_vectors_from_excel()

    # Step 2: Perform KMeans clustering with optimal cluster count
    clustered_results, cluster_centers, vectors, words = perform_kmeans_clustering(vocabulary_words, normalized_vectors)

    # Step 3: Load FastText model (backup, for potential similarity calculation)
    try:
        fasttext_model = load_model(fasttext_model_path)
    except Exception as e:
        error_message = str(e)[:100]
        print(f"Note: FastText model loading failed ({error_message}), but clustering results remain unaffected")
        fasttext_model = None

    # Step 4: Generate topics (sorted by similarity to cluster center)
    topic_results = generate_topic_keywords(clustered_results, cluster_centers, vectors, words, fasttext_model)

    # Step 5: Save topic results to Excel
    save_topic_results_to_excel(topic_results)